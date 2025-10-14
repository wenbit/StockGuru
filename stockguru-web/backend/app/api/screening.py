"""
筛选 API 路由
"""
from fastapi import APIRouter, HTTPException, BackgroundTasks
from fastapi.responses import HTMLResponse, Response
from typing import Optional
import logging
from datetime import datetime
from pathlib import Path
from jinja2 import Environment, FileSystemLoader
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

router = APIRouter()
logger = logging.getLogger(__name__)


class ScreeningRequest(BaseModel):
    """筛选请求"""
    date: str
    volume_top_n: Optional[int] = 100
    hot_top_n: Optional[int] = 100


class ScreeningResponse(BaseModel):
    """筛选响应"""
    task_id: str
    status: str
    message: str


@router.post("/screening", response_model=ScreeningResponse)
async def create_screening(request: ScreeningRequest, background_tasks: BackgroundTasks):
    """
    创建筛选任务
    
    - **date**: 筛选日期 (格式: YYYY-MM-DD)
    - **volume_top_n**: 成交量排名前N (默认: 100)
    - **hot_top_n**: 热度排名前N (默认: 100)
    """
    try:
        logger.info(f"创建筛选任务: date={request.date}")
        
        # 延迟导入服务
        from app.services.screening_service import ScreeningService
        screening_service = ScreeningService()
        
        # 创建任务
        result = await screening_service.create_screening_task(
            date=request.date,
            volume_top_n=request.volume_top_n,
            hot_top_n=request.hot_top_n
        )
        
        # 在后台执行筛选
        import asyncio
        
        def run_screening():
            """包装异步函数为同步函数"""
            asyncio.run(screening_service._execute_screening(
                result["task_id"],
                request.date,
                request.volume_top_n,
                request.hot_top_n
            ))
        
        background_tasks.add_task(run_screening)
        
        return result
        
    except Exception as e:
        logger.error(f"创建筛选任务失败: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/screening/{task_id}")
async def get_screening_result(task_id: str):
    """
    获取筛选结果
    
    - **task_id**: 任务ID
    """
    try:
        logger.info(f"获取筛选结果: task_id={task_id}")
        
        from app.services.screening_service import ScreeningService
        screening_service = ScreeningService()
        
        result = await screening_service.get_task_result(task_id)
        
        return result
        
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        logger.error(f"获取筛选结果失败: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/screening/{task_id}/export", response_class=HTMLResponse)
async def export_report(task_id: str):
    """
    导出筛选报告为 HTML
    
    Args:
        task_id: 任务ID
        
    Returns:
        HTML 格式的报告
    """
    try:
        logger.info(f"导出报告: {task_id}")
        
        # 获取任务结果
        from app.services.screening_service import ScreeningService
        screening_service = ScreeningService()
        result = await screening_service.get_task_result(task_id)
        
        if result['status'] != 'completed':
            raise HTTPException(status_code=400, detail="任务未完成，无法导出报告")
        
        if not result.get('results') or len(result['results']) == 0:
            raise HTTPException(status_code=404, detail="没有筛选结果")
        
        # 准备模板数据
        template_data = {
            'task_id': task_id,
            'date': datetime.now().strftime('%Y-%m-%d'),
            'result_count': len(result['results']),
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'results': result['results']
        }
        
        # 渲染模板
        template_dir = Path(__file__).parent.parent / 'templates'
        env = Environment(loader=FileSystemLoader(str(template_dir)))
        template = env.get_template('report_template.html')
        html_content = template.render(**template_data)
        
        logger.info(f"报告导出成功: {task_id}")
        return HTMLResponse(content=html_content)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出报告失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"导出报告失败: {str(e)}")


@router.get("/screening/{task_id}/export/excel")
async def export_report_excel(task_id: str):
    """
    导出筛选报告为 Excel 文件
    
    Args:
        task_id: 任务ID
        
    Returns:
        Excel 格式的报告
    """
    try:
        logger.info(f"导出Excel报告: {task_id}")
        
        # 获取任务结果
        from app.services.screening_service import ScreeningService
        screening_service = ScreeningService()
        result = await screening_service.get_task_result(task_id)
        
        if result['status'] != 'completed':
            raise HTTPException(status_code=400, detail="任务未完成，无法导出报告")
        
        if not result.get('results') or len(result['results']) == 0:
            raise HTTPException(status_code=404, detail="没有筛选结果")
        
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "筛选结果"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 15
        
        # 标题样式
        title_font = Font(name='微软雅黑', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        # 表头样式
        header_font = Font(name='微软雅黑', size=11, bold=True)
        header_fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 添加标题
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = '📈 StockGuru 股票筛选报告'
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = title_alignment
        title_cell.border = thin_border
        
        # 添加信息行
        ws.merge_cells('A2:H2')
        info_cell = ws['A2']
        info_cell.value = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  筛选结果: 共 {len(result['results'])} 只股票"
        info_cell.alignment = Alignment(horizontal='center')
        
        # 添加表头
        headers = ['排名', '股票代码', '股票名称', '动量分数', '综合评分', '最新价', '涨跌幅', '成交量']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 添加数据
        for idx, stock in enumerate(result['results'], 1):
            row = idx + 3
            
            # 数据
            ws.cell(row=row, column=1, value=stock.get('final_rank', idx))
            ws.cell(row=row, column=2, value=stock.get('stock_code', '-'))
            ws.cell(row=row, column=3, value=stock.get('stock_name', '-'))
            ws.cell(row=row, column=4, value=round(stock.get('momentum_score', 0), 2))
            ws.cell(row=row, column=5, value=round(stock.get('comprehensive_score', 0), 4))
            ws.cell(row=row, column=6, value=round(stock.get('close_price', 0), 2))
            ws.cell(row=row, column=7, value=round(stock.get('change_pct', 0), 2))
            ws.cell(row=row, column=8, value=int(stock.get('volume', 0)))
            
            # 设置样式
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # 涨跌幅颜色
                if col == 7:
                    change_pct = stock.get('change_pct', 0)
                    if change_pct > 0:
                        cell.font = Font(color='FF0000', bold=True)  # 红色
                    elif change_pct < 0:
                        cell.font = Font(color='00B050', bold=True)  # 绿色
        
        # 添加免责声明
        disclaimer_row = len(result['results']) + 5
        ws.merge_cells(f'A{disclaimer_row}:H{disclaimer_row}')
        disclaimer_cell = ws[f'A{disclaimer_row}']
        disclaimer_cell.value = '⚠️ 免责声明: 本报告仅供参考，不构成任何投资建议。市场有风险，投资需谨慎。'
        disclaimer_cell.font = Font(name='微软雅黑', size=9, italic=True)
        disclaimer_cell.alignment = Alignment(horizontal='center')
        
        # 保存到内存
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # 生成文件名
        filename = f"stockguru_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        logger.info(f"Excel报告导出成功: {task_id}")
        
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"导出Excel报告失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"导出Excel报告失败: {str(e)}")
